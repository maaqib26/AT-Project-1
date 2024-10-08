"""
excel_functions.py - Script to run excel function to extract test data from excel sheet
"""

import openpyxl
from openpyxl import load_workbook

class Excel_Operations:

    # Creating a constructor
    def __init__(self, file_name, sheet_number,):
        self.filename = file_name
        self.sheet_number = sheet_number

        # Method to get max row
    def row_count(self):
        workbook = load_workbook(self.filename)
        test_sheet = workbook[self.sheet_number]
        return test_sheet.max_row

    # Method to get max column
    def column_count(self):
        workbook = load_workbook(self.filename)
        test_sheet = workbook[self.sheet_number]
        return test_sheet.max_column

    # Method to read data from excel
    def read_data(self,row_number,column_number):
        workbook = load_workbook(self.filename)
        test_sheet = workbook[self.sheet_number]
        return test_sheet.cell(row=row_number,column=column_number).value

    # Method to write data in excel
    def write_data(self,row_number,column_number,data):
        workbook = load_workbook(self.filename)
        test_sheet = workbook[self.sheet_number]
        test_sheet.cell(row=row_number,column=column_number).value = data
        workbook.save(self.filename)



